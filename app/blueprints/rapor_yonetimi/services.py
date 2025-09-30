"""
Rapor işlemleri için servis modülü
Bu modül, raporlama ile ilgili iş mantığı işlemlerini gerçekleştirir.
"""
from datetime import datetime, timedelta
import os
import io
import json
import numpy as np
import pandas as pd
from flask import render_template, url_for, current_app
import weasyprint
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

from app.extensions import db
from app.blueprints.ogrenci_yonetimi.models import Ogrenci
from app.blueprints.ders_konu_yonetimi.models import Ders, Konu
from app.blueprints.calisma_programi.models import DersProgrami, DersIlerleme, KonuTakip
from app.blueprints.calisma_programi.services import ProgramService
from app.blueprints.deneme_sinavlari.models import DenemeSonuc
from app.blueprints.gorusme_defteri.models import GorusmeKaydi
# Devre dışı bırakılan modüller temizlendi
from app.blueprints.rapor_yonetimi.models import FaaliyetRaporu, RaporSablonu, IstatistikRaporu, RaporlananOlay

class RaporService:
    """
    Rapor işlemlerini yöneten servis sınıfı
    Gelişmiş metrikler ve analizler sunar
    """
    
    @staticmethod
    def create_faaliyet_raporu(baslik, donem, rapor_tipi, rapor_verileri, yorum=None):
        """
        Yeni faaliyet raporu oluştur
        
        Args:
            baslik: Rapor başlığı
            donem: Dönem bilgisi (2024-2025 Güz gibi)
            rapor_tipi: Rapor tipi (dönemsel, yıllık, aylık, haftalık)
            rapor_verileri: Rapor verileri (dict formatında)
            yorum: Rapor yorumu (opsiyonel)
            
        Returns:
            Dict: İşlem sonucunu ve oluşturulan rapor ID'sini içeren sözlük
        """
        try:
            # Önce rapor nesnesini oluştur
            rapor = FaaliyetRaporu(
                baslik=baslik,
                donem=donem,
                rapor_tipi=rapor_tipi,
                yorum=yorum,
                durum='taslak'
            )
            
            # JSON verilerini property üzerinden ayarla
            rapor.rapor_verileri_dict = rapor_verileri
            
            db.session.add(rapor)
            db.session.commit()
            
            return {
                'success': True,
                'message': 'Faaliyet raporu başarıyla oluşturuldu.',
                'rapor_id': rapor.id
            }
        except Exception as e:
            db.session.rollback()
            return {
                'success': False,
                'message': f"Rapor oluşturulurken hata oluştu: {str(e)}"
            }
    
    @staticmethod
    def generate_donemsel_rapor_verileri(baslangic_tarihi, bitis_tarihi):
        """
        Belirli bir dönem için rapor verilerini oluştur
        MEB E-Rehberlik formatına uygun
        
        Args:
            baslangic_tarihi: Dönem başlangıç tarihi (datetime.date)
            bitis_tarihi: Dönem bitiş tarihi (datetime.date)
            
        Returns:
            Dict: Rapor verilerini içeren sözlük
        """
        try:
            # Görüşme istatistiklerini hesapla
            gorusmeler = GorusmeKaydi.query.filter(
                GorusmeKaydi.tarih >= baslangic_tarihi,
                GorusmeKaydi.tarih <= bitis_tarihi
            ).all()
            
            # Öğrenci sayıları
            toplam_ogrenci = Ogrenci.query.count()
            
            # Anket ve ölçek istatistikleri - Devre dışı bırakıldı
            tamamlanan_anketler = 0  # Anket modülü kaldırıldı
            
            # Görüşme türlerine göre grupla
            bireysel_gorusme = sum(1 for g in gorusmeler if g.gorusulen_kisi == 'Öğrenci')
            veli_gorusme = sum(1 for g in gorusmeler if g.gorusulen_kisi in ['Veli', 'Anne', 'Baba', 'Vasi'])
            grup_gorusme = sum(1 for g in gorusmeler if g.gorusulen_kisi in ['Sınıf', 'Grup'])
            personel_gorusme = sum(1 for g in gorusmeler if g.gorusulen_kisi == 'Personel')
            
            # Aylara göre görüşme sayıları
            aylik_gorusmeler = {}
            for gorusme in gorusmeler:
                ay = gorusme.tarih.strftime('%Y-%m')
                ay_adi = gorusme.tarih.strftime('%B %Y')
                
                if ay not in aylik_gorusmeler:
                    aylik_gorusmeler[ay] = {
                        'ay_adi': ay_adi,
                        'toplam': 0,
                        'bireysel': 0,
                        'veli': 0,
                        'grup': 0,
                        'personel': 0
                    }
                
                aylik_gorusmeler[ay]['toplam'] += 1
                
                if gorusme.gorusulen_kisi == 'Öğrenci':
                    aylik_gorusmeler[ay]['bireysel'] += 1
                elif gorusme.gorusulen_kisi in ['Veli', 'Anne', 'Baba', 'Vasi']:
                    aylik_gorusmeler[ay]['veli'] += 1
                elif gorusme.gorusulen_kisi in ['Sınıf', 'Grup']:
                    aylik_gorusmeler[ay]['grup'] += 1
                elif gorusme.gorusulen_kisi == 'Personel':
                    aylik_gorusmeler[ay]['personel'] += 1
            
            # Sınıflara göre öğrenci ve görüşme dağılımı
            sinif_dagilimi = {}
            ogrenciler = Ogrenci.query.all()
            
            for ogrenci in ogrenciler:
                sinif = ogrenci.sinif
                
                if sinif not in sinif_dagilimi:
                    sinif_dagilimi[sinif] = {
                        'ogrenci_sayisi': 0,
                        'gorusme_sayisi': 0
                    }
                
                sinif_dagilimi[sinif]['ogrenci_sayisi'] += 1
                
                # Bu öğrenciyle yapılan görüşme sayısı
                ogrenci_gorusmeleri = sum(1 for g in gorusmeler if g.ogrenci_id == ogrenci.id)
                sinif_dagilimi[sinif]['gorusme_sayisi'] += ogrenci_gorusmeleri
            
            # Çalışma alanlarına göre görüşme dağılımı
            calisma_alanlari = {}
            for gorusme in gorusmeler:
                if not gorusme.calisma_alani:
                    continue
                
                alan = gorusme.calisma_alani
                
                if alan not in calisma_alanlari:
                    calisma_alanlari[alan] = 0
                
                calisma_alanlari[alan] += 1
            
            # Rapor verilerini oluştur (MEB E-Rehberlik formatında)
            rapor_verileri = {
                "okul_bilgileri": {
                    "okul_adi": "Uygulama kapsamında belirlenen okul adı",
                    "okul_turu": "Uygulama kapsamında belirlenen okul türü",
                    "il": "İl bilgisi",
                    "ilce": "İlçe bilgisi"
                },
                "rehberlik_kadrosu": {
                    "rehber_ogretmen_sayisi": 1,  # Sistem parametrelerinden alınabilir
                    "psikolojik_danismanlari": ["Rehber Öğretmen"]  # Sistem parametrelerinden alınabilir
                },
                "ozet_veriler": {
                    "rapor_tarih_araligi": f"{baslangic_tarihi.strftime('%d.%m.%Y')} - {bitis_tarihi.strftime('%d.%m.%Y')}",
                    "toplam_ogrenci": toplam_ogrenci,
                    "toplam_gorusme": len(gorusmeler),
                    "bireysel_gorusme": bireysel_gorusme,
                    "grup_gorusmesi": grup_gorusme,
                    "veli_gorusmesi": veli_gorusme,
                    "personel_gorusmesi": personel_gorusme,
                    "tamamlanan_anketler": tamamlanan_anketler
                },
                "aylik_istatistikler": [data for _, data in sorted(aylik_gorusmeler.items())],
                "sinif_dagilimi": [{"sinif": sinif, **veri} for sinif, veri in sinif_dagilimi.items()],
                "calisma_alanlari": [{"alan": alan, "gorusme_sayisi": sayi} for alan, sayi in calisma_alanlari.items()],
            }
            
            return rapor_verileri
            
        except Exception as e:
            current_app.logger.error(f"Dönemsel rapor verileri hesaplanırken hata oluştu: {str(e)}")
            return {
                "hata": f"Rapor verileri oluşturulurken bir hata oluştu: {str(e)}"
            }
    
    @staticmethod
    def generate_meb_faaliyet_raporu_pdf(rapor_id):
        """
        MEB formatında faaliyet raporu PDF'i oluştur
        
        Args:
            rapor_id: Faaliyet raporu ID
            
        Returns:
            Dict: İşlem sonucunu ve PDF dosya yolunu içeren sözlük
        """
        try:
            # Raporu veritabanından al
            rapor = FaaliyetRaporu.query.get(rapor_id)
            
            if not rapor:
                return {
                    'success': False,
                    'message': 'Rapor bulunamadı.'
                }
            
            # Rapor verilerini dict olarak al
            rapor_verileri = rapor.rapor_verileri_dict
            
            # Okul bilgilerini parametrelerden al
            from app.blueprints.parametre_yonetimi.services import SabitlerService
            okul_bilgisi = SabitlerService.get_okul_bilgisi()
            
            # Okul bilgilerini rapor verilerine ekle
            if not 'okul_bilgileri' in rapor_verileri:
                rapor_verileri['okul_bilgileri'] = {}
                
            if okul_bilgisi:
                rapor_verileri['okul_bilgileri']['okul_adi'] = okul_bilgisi.okul_adi
                rapor_verileri['okul_bilgileri']['okul_turu'] = "Anadolu Lisesi"  # Varsayılan olarak eklenebilir
                rapor_verileri['okul_bilgileri']['il'] = okul_bilgisi.il
                rapor_verileri['okul_bilgileri']['ilce'] = okul_bilgisi.ilce
                rapor_verileri['danisman_adi'] = okul_bilgisi.danisman_adi
            
            # HTML şablonunu render et
            html_content = render_template(
                'pdf/meb_faaliyet_raporu.html',
                rapor=rapor,
                rapor_verileri=rapor_verileri,
                tarih=datetime.now().strftime('%d.%m.%Y'),
                base_url=url_for('static', filename='', _external=True)
            )
            
            # Geçici PDF dosya yolu
            pdf_filename = f"meb_faaliyet_raporu_{rapor_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
            static_temp_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'static', 'temp')
            pdf_path = os.path.join(static_temp_dir, pdf_filename)
            os.makedirs(static_temp_dir, exist_ok=True)
            
            # PDF oluştur
            weasyprint.HTML(string=html_content).write_pdf(pdf_path)
            
            # Dosya URL'i
            pdf_url = url_for('static', filename=f"temp/{pdf_filename}")
            
            return {
                'success': True,
                'message': 'MEB faaliyet raporu PDF başarıyla oluşturuldu.',
                'pdf_path': pdf_path,
                'pdf_url': pdf_url
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f"PDF oluşturulurken hata oluştu: {str(e)}"
            }
    
    @staticmethod
    def export_faaliyet_raporu_excel(rapor_id):
        """
        Faaliyet raporunu Excel formatında dışa aktar
        
        Args:
            rapor_id: Faaliyet raporu ID
            
        Returns:
            BytesIO: Excel dosyasının içeriği
        """
        try:
            # Raporu veritabanından al
            rapor = FaaliyetRaporu.query.get(rapor_id)
            
            if not rapor:
                return None
            
            # Excel dosyasını oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Faaliyet Raporu"
            
            # Başlık stili
            baslik_font = Font(bold=True, size=14)
            baslik_alignment = Alignment(horizontal='center', vertical='center')
            baslik_fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
            baslik_border = Border(bottom=Side(style='thin'))
            
            # Altbaşlık stili
            altbaslik_font = Font(bold=True, size=12)
            
            # Tablo başlığı stili
            tablo_baslik_font = Font(bold=True, size=11)
            tablo_baslik_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Başlık
            ws.merge_cells('A1:G1')
            ws['A1'] = rapor.baslik
            ws['A1'].font = baslik_font
            ws['A1'].alignment = baslik_alignment
            
            # Dönem ve Tarih
            ws.merge_cells('A2:G2')
            ws['A2'] = f"Dönem: {rapor.donem} | Oluşturma Tarihi: {rapor.olusturma_tarihi.strftime('%d.%m.%Y')}"
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Property üzerinden rapor verilerini al
            rapor_data = rapor.rapor_verileri_dict
            
            # Okul Bilgileri
            row = 4
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "OKUL BİLGİLERİ"
            ws[f'A{row}'].font = altbaslik_font
            
            row += 1
            ws['A'+str(row)] = "Okul Adı"
            ws['B'+str(row)] = rapor_data.get("okul_bilgileri", {}).get("okul_adi", "")
            
            row += 1
            ws['A'+str(row)] = "Okul Türü"
            ws['B'+str(row)] = rapor_data.get("okul_bilgileri", {}).get("okul_turu", "")
            
            row += 1
            ws['A'+str(row)] = "İl / İlçe"
            ws['B'+str(row)] = f"{rapor_data.get('okul_bilgileri', {}).get('il', '')} / {rapor_data.get('okul_bilgileri', {}).get('ilce', '')}"
            
            # Rehberlik Kadrosu
            row += 2
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "REHBERLİK KADROSU"
            ws[f'A{row}'].font = altbaslik_font
            
            # Özet Veriler
            row += 2
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "ÖZET VERİLER"
            ws[f'A{row}'].font = altbaslik_font
            
            row += 1
            ws['A'+str(row)] = "Toplam Öğrenci"
            ws['B'+str(row)] = rapor_data.get("ozet_veriler", {}).get("toplam_ogrenci", 0)
            
            row += 1
            ws['A'+str(row)] = "Toplam Görüşme"
            ws['B'+str(row)] = rapor_data.get("ozet_veriler", {}).get("toplam_gorusme", 0)
            
            row += 1
            ws['A'+str(row)] = "Bireysel Görüşmeler"
            ws['B'+str(row)] = rapor_data.get("ozet_veriler", {}).get("bireysel_gorusme", 0)
            
            row += 1
            ws['A'+str(row)] = "Grup Görüşmeleri"
            ws['B'+str(row)] = rapor_data.get("ozet_veriler", {}).get("grup_gorusmesi", 0)
            
            row += 1
            ws['A'+str(row)] = "Veli Görüşmeleri"
            ws['B'+str(row)] = rapor_data.get("ozet_veriler", {}).get("veli_gorusmesi", 0)
            
            # Aylık İstatistikler
            row += 2
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "AYLIK İSTATİSTİKLER"
            ws[f'A{row}'].font = altbaslik_font
            
            row += 1
            aylik_istatistikler = rapor_data.get("aylik_istatistikler", [])
            
            if aylik_istatistikler:
                # Tablo başlıkları
                ws['A'+str(row)] = "Ay"
                ws['B'+str(row)] = "Toplam"
                ws['C'+str(row)] = "Bireysel"
                ws['D'+str(row)] = "Veli"
                ws['E'+str(row)] = "Grup"
                ws['F'+str(row)] = "Personel"
                
                # Başlık stilleri
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[col+str(row)].font = tablo_baslik_font
                    ws[col+str(row)].fill = tablo_baslik_fill
                
                # Veri girişi
                for ay_veri in aylik_istatistikler:
                    row += 1
                    ws['A'+str(row)] = ay_veri.get("ay_adi", "")
                    ws['B'+str(row)] = ay_veri.get("toplam", 0)
                    ws['C'+str(row)] = ay_veri.get("bireysel", 0)
                    ws['D'+str(row)] = ay_veri.get("veli", 0)
                    ws['E'+str(row)] = ay_veri.get("grup", 0)
                    ws['F'+str(row)] = ay_veri.get("personel", 0)
            
            # Sütun genişliklerini ayarla
            for i, column_cells in enumerate(ws.columns, 1):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws.column_dimensions[get_column_letter(i)].width = length + 4
            
            # Excel dosyasını BytesIO nesnesine yaz
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
            
        except Exception as e:
            current_app.logger.error(f"Excel dışa aktarma hatası: {str(e)}")
            return None
    
    @staticmethod
    def generate_istatistik_raporu(rapor_tipi, baslangic_tarihi, bitis_tarihi, filtreler=None):
        """
        İstatistiksel analiz raporu oluştur
        
        Args:
            rapor_tipi: Rapor tipi (analiz, karşılaştırma, performans, yönetim_özeti)
            baslangic_tarihi: Başlangıç tarihi
            bitis_tarihi: Bitiş tarihi
            filtreler: Filtre parametreleri (dict)
            
        Returns:
            Dict: İşlem sonucunu ve rapor verilerini içeren sözlük
        """
        try:
            rapor_verileri = {}
            grafikler = {}
            
            # Rapor tipine göre veri hazırla
            if rapor_tipi == 'analiz':
                # Görüşme analizleri
                gorusmeler = GorusmeKaydi.query.filter(
                    GorusmeKaydi.tarih >= baslangic_tarihi,
                    GorusmeKaydi.tarih <= bitis_tarihi
                ).all()
                
                # Görüşme türlerine göre dağılım
                gorusme_turleri = {
                    'Bireysel': sum(1 for g in gorusmeler if g.gorusulen_kisi == 'Öğrenci'),
                    'Veli': sum(1 for g in gorusmeler if g.gorusulen_kisi in ['Veli', 'Anne', 'Baba', 'Vasi']),
                    'Grup': sum(1 for g in gorusmeler if g.gorusulen_kisi in ['Sınıf', 'Grup']),
                    'Personel': sum(1 for g in gorusmeler if g.gorusulen_kisi == 'Personel')
                }
                
                # Çalışma alanlarına göre dağılım
                calisma_alanlari = {}
                for gorusme in gorusmeler:
                    if not gorusme.calisma_alani:
                        continue
                    
                    alan = gorusme.calisma_alani
                    
                    if alan not in calisma_alanlari:
                        calisma_alanlari[alan] = 0
                    
                    calisma_alanlari[alan] += 1
                
                # Grafik verilerini hazırla
                gorusme_turleri_grafik = {
                    'labels': list(gorusme_turleri.keys()),
                    'values': list(gorusme_turleri.values()),
                    'colors': ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e']
                }
                
                calisma_alanlari_grafik = {
                    'labels': list(calisma_alanlari.keys()),
                    'values': list(calisma_alanlari.values()),
                    'colors': ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', '#5a5c69'][:len(calisma_alanlari)]
                }
                
                # Rapor verilerini hazırla
                rapor_verileri = {
                    'baslangic_tarihi': baslangic_tarihi.strftime('%d.%m.%Y'),
                    'bitis_tarihi': bitis_tarihi.strftime('%d.%m.%Y'),
                    'toplam_gorusme': len(gorusmeler),
                    'gorusme_turleri': gorusme_turleri,
                    'calisma_alanlari': calisma_alanlari
                }
                
                # Grafikleri hazırla
                grafikler = {
                    'gorusme_turleri': gorusme_turleri_grafik,
                    'calisma_alanlari': calisma_alanlari_grafik
                }
                
            elif rapor_tipi == 'karşılaştırma':
                # Öğrenci grupları arası karşılaştırma
                
                # Öğrenci grupları yerine sınıf tabanlı gruplandırma kullanılacak
                grup_karsilastirma = []
                
                # Sınıflara göre gruplandırma yapalım (öğrenci grupları yerine)
                siniflar = db.session.query(Ogrenci.sinif).distinct().all()
                
                # Her bir sınıfı bir grup olarak değerlendirelim
                for sinif_row in siniflar:
                    sinif = sinif_row[0]
                    if sinif is None or sinif == '':
                        continue  # Sınıf bilgisi olmayan öğrencileri atla
                        
                    # Sınıftaki öğrencileri bul
                    ogrenci_ids = [o.id for o in Ogrenci.query.filter_by(sinif=sinif).all()]
                    
                    # Grup için istatistikler
                    gorusme_sayisi = GorusmeKaydi.query.filter(
                        GorusmeKaydi.ogrenci_id.in_(ogrenci_ids),
                        GorusmeKaydi.tarih >= baslangic_tarihi,
                        GorusmeKaydi.tarih <= bitis_tarihi
                    ).count()
                    
                    deneme_ortalama = db.session.query(func.avg(DenemeSonuc.puan_tyt)).filter(
                        DenemeSonuc.ogrenci_id.in_(ogrenci_ids),
                        DenemeSonuc.tarih >= baslangic_tarihi,
                        DenemeSonuc.tarih <= bitis_tarihi
                    ).scalar() or 0
                    
                    grup_karsilastirma.append({
                        'grup_adi': f"Sınıf {sinif}",  # Sınıf adını kullan
                        'ogrenci_sayisi': len(ogrenci_ids),
                        'gorusme_sayisi': gorusme_sayisi,
                        'deneme_ortalama': round(float(deneme_ortalama), 2)
                    })
                
                # Grafik verisi
                grup_karsilastirma_grafik = {
                    'labels': [grup['grup_adi'] for grup in grup_karsilastirma],
                    'datasets': [
                        {
                            'label': 'Öğrenci Sayısı',
                            'data': [grup['ogrenci_sayisi'] for grup in grup_karsilastirma],
                            'backgroundColor': '#4e73df'
                        },
                        {
                            'label': 'Görüşme Sayısı',
                            'data': [grup['gorusme_sayisi'] for grup in grup_karsilastirma],
                            'backgroundColor': '#1cc88a'
                        },
                        {
                            'label': 'Deneme Ortalaması',
                            'data': [grup['deneme_ortalama'] for grup in grup_karsilastirma],
                            'backgroundColor': '#f6c23e'
                        }
                    ]
                }
                
                # Rapor verilerini hazırla
                rapor_verileri = {
                    'baslangic_tarihi': baslangic_tarihi.strftime('%d.%m.%Y'),
                    'bitis_tarihi': bitis_tarihi.strftime('%d.%m.%Y'),
                    'gruplar': grup_karsilastirma
                }
                
                # Grafikleri hazırla
                grafikler = {
                    'grup_karsilastirma': grup_karsilastirma_grafik
                }
                
            elif rapor_tipi == 'performans':
                # Rehberlik performans göstergeleri
                
                # Tarih aralığındaki günler
                tarih_farki = (bitis_tarihi - baslangic_tarihi).days + 1
                hafta_sayisi = tarih_farki // 7 + (1 if tarih_farki % 7 > 0 else 0)
                
                # Görüşme sayıları
                toplam_gorusme = GorusmeKaydi.query.filter(
                    GorusmeKaydi.tarih >= baslangic_tarihi,
                    GorusmeKaydi.tarih <= bitis_tarihi
                ).count()
                
                # Anket ve ölçek sayıları - Devre dışı bırakıldı
                tamamlanan_anketler = 0  # Anket modülü kaldırıldı
                
                # Toplam öğrenci sayısı
                toplam_ogrenci = Ogrenci.query.count()
                
                # Performans göstergeleri
                gunluk_gorusme = round(toplam_gorusme / tarih_farki, 2) if tarih_farki > 0 else 0
                haftalik_gorusme = round(toplam_gorusme / hafta_sayisi, 2) if hafta_sayisi > 0 else 0
                ogrenci_basina_gorusme = round(toplam_gorusme / toplam_ogrenci, 2) if toplam_ogrenci > 0 else 0
                
                # Önceki dönemle karşılaştırma (örnek)
                onceki_baslangic = baslangic_tarihi - timedelta(days=tarih_farki)
                onceki_bitis = baslangic_tarihi - timedelta(days=1)
                
                onceki_gorusme = GorusmeKaydi.query.filter(
                    GorusmeKaydi.tarih >= onceki_baslangic,
                    GorusmeKaydi.tarih <= onceki_bitis
                ).count()
                
                gorusme_degisim = round(((toplam_gorusme - onceki_gorusme) / onceki_gorusme * 100), 2) if onceki_gorusme > 0 else 100
                
                # Rapor verilerini hazırla
                rapor_verileri = {
                    'baslangic_tarihi': baslangic_tarihi.strftime('%d.%m.%Y'),
                    'bitis_tarihi': bitis_tarihi.strftime('%d.%m.%Y'),
                    'toplam_gorusme': toplam_gorusme,
                    'tamamlanan_anketler': tamamlanan_anketler,
                    'toplam_ogrenci': toplam_ogrenci,
                    'gun_sayisi': tarih_farki,
                    'hafta_sayisi': hafta_sayisi,
                    'gunluk_gorusme': gunluk_gorusme,
                    'haftalik_gorusme': haftalik_gorusme,
                    'ogrenci_basina_gorusme': ogrenci_basina_gorusme,
                    'onceki_donem_gorusme': onceki_gorusme,
                    'gorusme_degisim': gorusme_degisim
                }
                
                # Performans göstergeleri grafikleri
                performans_grafik = {
                    'labels': ['Günlük Görüşme', 'Haftalık Görüşme', 'Öğrenci Başına Görüşme'],
                    'values': [gunluk_gorusme, haftalik_gorusme, ogrenci_basina_gorusme],
                    'colors': ['#4e73df', '#1cc88a', '#36b9cc']
                }
                
                degisim_grafik = {
                    'labels': ['Önceki Dönem', 'Şu Anki Dönem'],
                    'values': [onceki_gorusme, toplam_gorusme],
                    'colors': ['#e74a3b', '#4e73df']
                }
                
                # Grafikleri hazırla
                grafikler = {
                    'performans': performans_grafik,
                    'degisim': degisim_grafik
                }
                
            elif rapor_tipi == 'yönetim_özeti':
                # Okul yönetimine sunulacak özet rapor
                
                # Görüşme istatistikleri
                gorusmeler = GorusmeKaydi.query.filter(
                    GorusmeKaydi.tarih >= baslangic_tarihi,
                    GorusmeKaydi.tarih <= bitis_tarihi
                ).all()
                
                # Aylara göre görüşme sayıları
                aylik_gorusmeler = {}
                for gorusme in gorusmeler:
                    ay = gorusme.tarih.strftime('%Y-%m')
                    ay_adi = gorusme.tarih.strftime('%B %Y')
                    
                    if ay not in aylik_gorusmeler:
                        aylik_gorusmeler[ay] = {
                            'ay_adi': ay_adi,
                            'toplam': 0
                        }
                    
                    aylik_gorusmeler[ay]['toplam'] += 1
                
                # Sınıf bazında öğrenci gelişim durumu
                sinif_gelisim = {}
                ogrenciler = Ogrenci.query.all()
                
                for ogrenci in ogrenciler:
                    sinif = ogrenci.sinif
                    
                    if sinif not in sinif_gelisim:
                        sinif_gelisim[sinif] = {
                            'ogrenci_sayisi': 0,
                            'ortalama_ilerleme': 0,
                            'gorusme_sayisi': 0,
                            'risk_durumu': {'yüksek': 0, 'orta': 0, 'düşük': 0}
                        }
                    
                    sinif_gelisim[sinif]['ogrenci_sayisi'] += 1
                    
                    # Bu öğrencinin ders ilerlemelerini bul
                    ilerlemeler = DersIlerleme.query.filter_by(ogrenci_id=ogrenci.id).all()
                    ortalama_ilerleme = sum(i.tamamlama_yuzdesi for i in ilerlemeler) / len(ilerlemeler) if ilerlemeler else 0
                    sinif_gelisim[sinif]['ortalama_ilerleme'] += ortalama_ilerleme
                    
                    # Bu öğrenciyle yapılan görüşme sayısı
                    ogrenci_gorusmeleri = sum(1 for g in gorusmeler if g.ogrenci_id == ogrenci.id)
                    sinif_gelisim[sinif]['gorusme_sayisi'] += ogrenci_gorusmeleri
                    
                    # Risk durumu (örnek hesaplama)
                    risk_seviyesi = 'düşük'
                    if ortalama_ilerleme < 30:
                        risk_seviyesi = 'yüksek'
                    elif ortalama_ilerleme < 60:
                        risk_seviyesi = 'orta'
                    
                    sinif_gelisim[sinif]['risk_durumu'][risk_seviyesi] += 1
                
                # Ortalama ilerleme hesapla
                for sinif in sinif_gelisim:
                    sinif_gelisim[sinif]['ortalama_ilerleme'] = round(
                        sinif_gelisim[sinif]['ortalama_ilerleme'] / sinif_gelisim[sinif]['ogrenci_sayisi'], 2
                    ) if sinif_gelisim[sinif]['ogrenci_sayisi'] > 0 else 0
                
                # Risk durumu özeti
                risk_durumu_ozet = {
                    'yüksek': sum(s['risk_durumu']['yüksek'] for s in sinif_gelisim.values()),
                    'orta': sum(s['risk_durumu']['orta'] for s in sinif_gelisim.values()),
                    'düşük': sum(s['risk_durumu']['düşük'] for s in sinif_gelisim.values())
                }
                
                # Grafik verileri
                aylik_trend_grafik = {
                    'labels': [data['ay_adi'] for _, data in sorted(aylik_gorusmeler.items())],
                    'values': [data['toplam'] for _, data in sorted(aylik_gorusmeler.items())],
                    'color': '#4e73df'
                }
                
                sinif_ilerleme_grafik = {
                    'labels': list(sinif_gelisim.keys()),
                    'values': [sinif['ortalama_ilerleme'] for sinif in sinif_gelisim.values()],
                    'colors': ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b'][:len(sinif_gelisim)]
                }
                
                risk_durumu_grafik = {
                    'labels': ['Yüksek Risk', 'Orta Risk', 'Düşük Risk'],
                    'values': [risk_durumu_ozet['yüksek'], risk_durumu_ozet['orta'], risk_durumu_ozet['düşük']],
                    'colors': ['#e74a3b', '#f6c23e', '#1cc88a']
                }
                
                # Rapor verilerini hazırla
                rapor_verileri = {
                    'baslangic_tarihi': baslangic_tarihi.strftime('%d.%m.%Y'),
                    'bitis_tarihi': bitis_tarihi.strftime('%d.%m.%Y'),
                    'toplam_gorusme': len(gorusmeler),
                    'aylik_gorusmeler': [data for _, data in sorted(aylik_gorusmeler.items())],
                    'sinif_gelisim': [{'sinif': sinif, **veri} for sinif, veri in sinif_gelisim.items()],
                    'risk_durumu_ozet': risk_durumu_ozet
                }
                
                # Grafikleri hazırla
                grafikler = {
                    'aylik_trend': aylik_trend_grafik,
                    'sinif_ilerleme': sinif_ilerleme_grafik,
                    'risk_durumu': risk_durumu_grafik
                }
            
            # Yeni rapor kaydı oluştur
            baslik = f"{rapor_tipi.capitalize()} Raporu ({baslangic_tarihi.strftime('%d.%m.%Y')} - {bitis_tarihi.strftime('%d.%m.%Y')})"
            
            rapor = IstatistikRaporu(
                baslik=baslik,
                rapor_tipi=rapor_tipi,
                baslangic_tarihi=baslangic_tarihi,
                bitis_tarihi=bitis_tarihi
            )
            rapor.filtreler_dict = filtreler
            rapor.rapor_verileri_dict = rapor_verileri
            rapor.grafikler_dict = grafikler
            
            db.session.add(rapor)
            db.session.commit()
            
            return {
                'success': True,
                'message': 'İstatistiksel rapor başarıyla oluşturuldu.',
                'rapor_id': rapor.id,
                'rapor_verileri': rapor_verileri,
                'grafikler': grafikler
            }
            
        except Exception as e:
            db.session.rollback()
            return {
                'success': False,
                'message': f"Rapor oluşturulurken hata oluştu: {str(e)}"
            }
    
    @staticmethod
    def generate_haftalik_plan_pdf(ogrenci_id, baslangic_tarihi=None):
        """
        Haftalık ders planı PDF raporu oluştur
        
        Args:
            ogrenci_id: Öğrenci ID
            baslangic_tarihi: Hafta başlangıç tarihi (opsiyonel, varsayılan bugün)
            
        Returns:
            Dict: İşlem sonucunu ve PDF dosya yolunu içeren sözlük
        """
        # Öğrenci kontrolü
        ogrenci = Ogrenci.query.get(ogrenci_id)
        if not ogrenci:
            return {
                'success': False, 
                'message': 'Öğrenci bulunamadı.'
            }
        
        # Haftalık plan verilerini al
        haftalik_plan = ProgramService.generate_haftalik_plan(ogrenci_id, baslangic_tarihi)
        if not haftalik_plan['success']:
            return haftalik_plan
        
        try:
            # HTML şablonunu render et
            html_content = render_template(
                'pdf/haftalik_plan_pdf.html',
                ogrenci=haftalik_plan['ogrenci'],
                hafta=haftalik_plan['hafta'],
                gunler=haftalik_plan['gunler'],
                ders_ilerleme=haftalik_plan['ders_ilerleme'],
                tarih=datetime.now().strftime('%d.%m.%Y'),
                base_url=url_for('static', filename='', _external=True)
            )
            
            # Geçici PDF dosya yolu
            pdf_filename = f"haftalik_plan_{ogrenci_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
            pdf_path = os.path.join('/home/runner/workspace/static/temp', pdf_filename)
            os.makedirs('/home/runner/workspace/static/temp', exist_ok=True)
            
            # PDF oluştur
            weasyprint.HTML(string=html_content).write_pdf(pdf_path)
            
            # Dosya URL'i
            pdf_url = url_for('static', filename=f"temp/{pdf_filename}")
            
            return {
                'success': True,
                'message': 'Haftalık plan PDF raporu başarıyla oluşturuldu.',
                'pdf_path': pdf_path,
                'pdf_url': pdf_url
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f"PDF oluşturulurken hata oluştu: {str(e)}"
            }
    
    @staticmethod
    def generate_konu_plani_pdf(ogrenci_id, ders_id):
        """
        Konu planı PDF raporu oluştur
        
        Args:
            ogrenci_id: Öğrenci ID
            ders_id: Ders ID
            
        Returns:
            Dict: İşlem sonucunu ve PDF dosya yolunu içeren sözlük
        """
        # Öğrenci kontrolü
        ogrenci = Ogrenci.query.get(ogrenci_id)
        if not ogrenci:
            return {
                'success': False, 
                'message': 'Öğrenci bulunamadı.'
            }
        
        # Ders kontrolü
        ders = Ders.query.get(ders_id)
        if not ders:
            return {
                'success': False, 
                'message': 'Ders bulunamadı.'
            }
        
        # Ders konuları
        konular = Konu.query.filter_by(ders_id=ders_id).order_by(Konu.sira).all()
        
        # Konu takip bilgileri
        konu_takip_map = {}
        for konu in konular:
            takip = KonuTakip.query.filter_by(ogrenci_id=ogrenci_id, konu_id=konu.id).first()
            if takip:
                konu_takip_map[konu.id] = takip
        
        # Ders ilerleme bilgisi
        ilerleme = DersIlerleme.query.filter_by(ogrenci_id=ogrenci_id, ders_id=ders_id).first()
        
        try:
            # HTML şablonunu render et
            html_content = render_template(
                'pdf/konu_plani_pdf.html',
                ogrenci=ogrenci,
                ders=ders,
                konular=konular,
                konu_takip_map=konu_takip_map,
                ilerleme=ilerleme,
                tarih=datetime.now().strftime('%d.%m.%Y'),
                base_url=url_for('static', filename='', _external=True)
            )
            
            # Geçici PDF dosya yolu
            pdf_filename = f"konu_plani_{ogrenci_id}_{ders_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
            pdf_path = os.path.join('/home/runner/workspace/static/temp', pdf_filename)
            os.makedirs('/home/runner/workspace/static/temp', exist_ok=True)
            
            # PDF oluştur
            weasyprint.HTML(string=html_content).write_pdf(pdf_path)
            
            # Dosya URL'i
            pdf_url = url_for('static', filename=f"temp/{pdf_filename}")
            
            return {
                'success': True,
                'message': 'Konu planı PDF raporu başarıyla oluşturuldu.',
                'pdf_path': pdf_path,
                'pdf_url': pdf_url
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f"PDF oluşturulurken hata oluştu: {str(e)}"
            }
    
    @staticmethod
    def generate_ilerleme_raporu_pdf(ogrenci_id):
        """
        İlerleme raporu PDF raporu oluştur
        
        Args:
            ogrenci_id: Öğrenci ID
            
        Returns:
            Dict: İşlem sonucunu ve PDF dosya yolunu içeren sözlük
        """
        # Öğrenci kontrolü
        ogrenci = Ogrenci.query.get(ogrenci_id)
        if not ogrenci:
            return {
                'success': False, 
                'message': 'Öğrenci bulunamadı.'
            }
        
        # Tamamlanma tarihlerini güncelle
        from app.blueprints.calisma_programi.completion_calculator import update_ders_completion_dates
        completion_data = update_ders_completion_dates(ogrenci_id)
        
        # Ders ilerleme bilgileri
        ilerlemeler = DersIlerleme.query.filter_by(ogrenci_id=ogrenci_id).all()
        
        # İlerleme verileri
        ilerleme_verileri = []
        for ilerleme in ilerlemeler:
            ders = Ders.query.get(ilerleme.ders_id)
            
            # Konuları al
            konular = Konu.query.filter_by(ders_id=ders.id).all()
            konu_sayisi = len(konular)
            
            # Tamamlanan konuları say
            tamamlanan = 0
            for konu in konular:
                takip = KonuTakip.query.filter_by(ogrenci_id=ogrenci_id, konu_id=konu.id).first()
                if takip and takip.tamamlandi:
                    tamamlanan += 1
            
            # Ders için haftalık çalışma süresini hesapla
            from app.blueprints.calisma_programi.models import DersProgrami
            ders_programlari = DersProgrami.query.filter_by(
                ogrenci_id=ogrenci_id,
                ders_id=ders.id
            ).all()
            haftalik_sure = sum(dp.sure_dakika() for dp in ders_programlari)
            
            # Veri hazırla
            ilerleme_verileri.append({
                'ders': ders,
                'ilerleme': ilerleme,
                'konu_sayisi': konu_sayisi,
                'tamamlanan': tamamlanan,
                'kalan': konu_sayisi - tamamlanan,
                'haftalik_sure': haftalik_sure
            })
        
        # İlerleme verilerini tamamlama yüzdesine göre sırala
        ilerleme_verileri.sort(key=lambda x: x['ilerleme'].tamamlama_yuzdesi, reverse=True)
        
        try:
            # HTML şablonunu render et
            now = datetime.now().date()
            html_content = render_template(
                'pdf/ilerleme_raporu_pdf.html',
                ogrenci=ogrenci,
                ilerleme_verileri=ilerleme_verileri,
                tarih=datetime.now().strftime('%d.%m.%Y'),
                now=now,
                konu_takipleri=[],  # Bu veriler raporda kullanılabilir ancak şu an kapsam dışı
                deneme_sonuclari=[],  # Bu veriler raporda kullanılabilir ancak şu an kapsam dışı
                en_gec_tamamlanma_tarihi=completion_data.get('en_gec_tamamlanma_tarihi'),
                base_url=url_for('static', filename='', _external=True)
            )
            
            # Geçici PDF dosya yolu
            pdf_filename = f"ilerleme_raporu_{ogrenci_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
            # Uygulama kök dizininden itibaren static/temp yolunu kullan
            static_temp_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'static', 'temp')
            pdf_path = os.path.join(static_temp_dir, pdf_filename)
            # Dizinin varlığından emin olalım
            os.makedirs(static_temp_dir, exist_ok=True)
            
            # PDF oluştur
            weasyprint.HTML(string=html_content).write_pdf(pdf_path)
            
            # Dosya URL'i
            pdf_url = url_for('static', filename=f"temp/{pdf_filename}")
            
            return {
                'success': True,
                'message': 'İlerleme raporu PDF raporu başarıyla oluşturuldu.',
                'pdf_path': pdf_path,
                'pdf_url': pdf_url
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f"PDF oluşturulurken hata oluştu: {str(e)}"
            }